import streamlit as st
import os
import openpyxl
import json
import re
import time
import gspread
from google import genai
from dotenv import load_dotenv
from google.oauth2.service_account import Credentials
from sheet_utils import (
    parse_application_id,
    normalize_record_keys,
    find_row_by_application_id,
    extract_medical_histories,
    extract_lifestyle_habits,
)

# 載入環境變數
load_dotenv()

def build_language_system_rule(lang: str, word_limit: int) -> str:
    return f"""
# LANGUAGE CONSTRAINT — ABSOLUTE RULE (HIGHEST PRIORITY)

The user has selected the output language: {lang}

You MUST write the ENTIRE response strictly in this language.
Any violation makes the response INVALID.
You MUST keep the total output within {word_limit} characters (non-space) for the JSON values.

- If lang is "English":
  - Respond in English ONLY
  - DO NOT output any Chinese characters (no 中文/漢字)
- If lang is "繁體中文":
  - Respond in Traditional Chinese ONLY
- If lang is "日本語":
  - すべて日本語で回答してください
- If lang is "한국어":
  - 모든 내용을 한국어로 작성하세요
- If lang is "Tiếng Việt":
  - Trả lời hoàn toàn bằng tiếng Việt

Return JSON ONLY. No extra text outside JSON.
""".strip()

def is_language_valid(text: str, lang: str) -> bool:
    if lang == "English":
        return not re.search(r"[\u4e00-\u9fff\u3040-\u30ff]", text)
    if lang == "繁體中文":
        return not re.search(r"[\u3040-\u30ff]", text)
    if lang == "日本語":
        return bool(re.search(r"[\u3040-\u30ff]", text))
    if lang == "한국어":
        return bool(re.search(r"[\uac00-\ud7af]", text))
    if lang == "Tiếng Việt":
        return bool(re.search(r"[A-Za-zÀ-ỹ]", text))
    return True

def count_output_length(text: str, lang: str) -> int:
    return len(re.findall(r"\S", text))

def normalize_report_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, dict):
        if not value:
            return ""
        return " ".join(str(v) for v in value.values())
    if isinstance(value, list):
        if not value:
            return ""
        return " ".join(str(v) for v in value)
    return str(value)

def min_section_length(word_limit: int) -> int:
    return max(20, int(word_limit * 0.03))

def validate_report_output(report: dict, lang: str, word_limit: int) -> tuple[bool, str, int]:
    combined_text = " ".join(normalize_report_value(v) for v in report.values())
    if not is_language_valid(combined_text, lang):
        return False, "語言不符合選擇", count_output_length(combined_text, lang)
    section_min = min_section_length(word_limit)
    required_keys = ["maintenance", "tracking", "nutrition", "supplements", "lifestyle"]
    for key in required_keys:
        section_text = normalize_report_value(report.get(key)).strip()
        if not section_text:
            return False, f"{key} 欄位內容為空", count_output_length(combined_text, lang)
        section_length = count_output_length(section_text, lang)
        if section_length < section_min:
            return False, f"{key} 欄位內容過短", count_output_length(combined_text, lang)
    length = count_output_length(combined_text, lang)
    if length > word_limit:
        return False, f"超過字數限制（{length}/{word_limit}）", length
    return True, "", length

def build_length_budget(word_limit: int) -> dict:
    weights = {
        "maintenance": 0.2,
        "tracking": 0.15,
        "nutrition": 0.2,
        "supplements": 0.2,
        "lifestyle": 0.25,
    }
    remaining = word_limit
    budget = {}
    ordered_keys = list(weights.keys())
    for key in ordered_keys[:-1]:
        allocation = max(1, int(word_limit * weights[key]))
        allocation = min(allocation, remaining)
        budget[key] = allocation
        remaining -= allocation
    budget[ordered_keys[-1]] = max(1, remaining)
    return budget

def format_budget_hint(budget: dict) -> str:
    return (
        f'maintenance≤{budget["maintenance"]}, '
        f'tracking≤{budget["tracking"]}, '
        f'nutrition≤{budget["nutrition"]}, '
        f'supplements≤{budget["supplements"]}, '
        f'lifestyle≤{budget["lifestyle"]}'
    )


def load_records_from_google_sheet(sheet_url: str, worksheet_name: str | None = None, worksheet_gid: int | None = None):
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly",
    ]

    service_account_info = None
    if "gcp_service_account" in st.secrets:
        service_account_info = dict(st.secrets["gcp_service_account"])
    else:
        service_account_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
        if service_account_json:
            service_account_info = json.loads(service_account_json)

    if service_account_info:
        credentials = Credentials.from_service_account_info(service_account_info, scopes=scopes)
    else:
        service_account_file = os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE", "").strip()
        if not service_account_file:
            raise ValueError("缺少 Google Service Account 設定，請設定 Streamlit secrets 或 GOOGLE_SERVICE_ACCOUNT_FILE / GOOGLE_SERVICE_ACCOUNT_JSON。")
        credentials = Credentials.from_service_account_file(service_account_file, scopes=scopes)

    gc = gspread.authorize(credentials)
    spreadsheet = gc.open_by_url(sheet_url)
    if worksheet_gid is not None:
        worksheet = spreadsheet.get_worksheet_by_id(worksheet_gid)
    elif worksheet_name:
        worksheet = spreadsheet.worksheet(worksheet_name)
    else:
        worksheet = spreadsheet.sheet1
    return normalize_record_keys(worksheet.get_all_records())

# --- 1. 核心邏輯：擷取 Excel 數據 ---
def extract_data_from_upload(uploaded_file, threshold_low=30, threshold_std=37):
    # Streamlit 上傳的檔案是 BytesIO 物件
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    
    # 版型判定
    count_a = sum(1 for r in range(3, 15) if ws.cell(row=r, column=1).value)
    count_b = sum(1 for r in range(3, 15) if ws.cell(row=r, column=2).value)
    is_5_slot = count_b >= count_a * 1.2

    user_info = {}
    if is_5_slot:
        user_info['age'] = ws.cell(row=2, column=5).value
        user_info['gender'] = ws.cell(row=2, column=6).value
        start_row, step, p_col = 3, 5, 2
    else:
        user_info['age'] = ws.cell(row=2, column=7).value
        user_info['gender'] = ws.cell(row=2, column=8).value
        start_row, step, p_col = 2, 3, 1

    all_scored_items = []
    for row in range(start_row, ws.max_row + 1, step):
        p_name = ws.cell(row=row, column=p_col).value
        score_val = ws.cell(row=row, column=10).value
        if p_name and score_val is not None:
            try:
                all_scored_items.append({"name": str(p_name), "score": float(score_val)})
            except: continue

    # 階層式篩選
    tier_1 = [item['name'] for item in all_scored_items if item['score'] < threshold_low]
    if tier_1:
        return user_info, tier_1, "極低分 (<30)"
    
    tier_2 = [item['name'] for item in all_scored_items if item['score'] < threshold_std]
    return user_info, tier_2, "標準篩選 (<37)"

# --- 2. 格式化工具 ---
def format_output(content):
    if isinstance(content, list):
        lines = []
        for idx, entry in enumerate(content, 1):
            if isinstance(entry, dict):
                val_str = " ".join([str(v) for v in entry.values()])
                lines.append(f"{idx}. {val_str}")
            else:
                lines.append(f"{idx}. {entry}")
        return "\n".join(lines)
    return str(content).strip()

# --- 3. Streamlit 網頁介面 ---
st.set_page_config(page_title="AI 營養報告生成器", layout="wide")
st.title("🧬 印度AI 細胞解碼報告生成器")

with st.sidebar:
    st.header("⚙️ 參數設定")
    # API Key 優先讀取 Secrets，若無則顯示輸入框
    api_key_val = os.getenv("GEMINI_API_KEY", "")
    api_key = st.text_input("Gemini API Key", type="password", value=api_key_val)
    lang = st.selectbox("輸出語言", ["繁體中文", "English", "日本語", "한국어", "Tiếng Việt"], index=0)
    word_limit = st.number_input("字數限制", value=800)

# 【修改點 1】：移除提示詞上傳區，僅保留 Excel 上傳
up_excel = st.file_uploader("上傳檢測 Excel 檔案", type=["xlsx"])

# 固定設定：Google Sheet 與提示詞檔
GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1JDaap1KOnKn4ZefISp27edfW1nWJyf4EFWWrd4dxVdU/edit?resourcekey=&gid=1866179831#gid=1866179831"
GOOGLE_SHEET_WORKSHEET = ""
GOOGLE_SHEET_GID = 1866179831
PROMPT_FILE_NAME = "系統提示詞_v3.1_純文字.txt"

if st.button("🚀 開始分析報告") and up_excel and api_key:
    # 檢查提示詞檔案是否存在
    if not os.path.exists(PROMPT_FILE_NAME):
        st.error(f"❌ 找不到設定檔：{PROMPT_FILE_NAME}。請確認檔案已上傳至 GitHub。")
    else:
        try:
            client = genai.Client(api_key=api_key)
            
            # 【修改點 3】：自動讀取本地檔案中的提示詞
            with open(PROMPT_FILE_NAME, "r", encoding="utf-8") as f:
                bg_prompt = f.read()
        
            with st.spinner("正在逐項分析中，請稍候..."):
                user_info, items, mode = extract_data_from_upload(up_excel)

                # 解析申請單編號（檔名格式不符時給出警告，繼續執行）
                try:
                    application_id = parse_application_id(up_excel.name)
                except ValueError as e:
                    application_id = ""
                    st.warning(f"⚠️ 無法從檔名解析申請單編號：{e}（病史將顯示為未提供）")

                # 從 Google Sheet 讀取資料
                records = load_records_from_google_sheet(GOOGLE_SHEET_URL, GOOGLE_SHEET_WORKSHEET or None, GOOGLE_SHEET_GID)

                # ===== 診斷輸出（debug，確認後可移除）=====
                st.write(f"🔍 DEBUG: 共讀取 {len(records)} 筆記錄")
                if records:
                    st.write(f"🔍 DEBUG: 欄位名稱 = {list(records[0].keys())}")
                # ===== 診斷輸出結束 =====

                # 找對應資料列（找不到時顯示警告，繼續執行）
                matched_row = find_row_by_application_id(records, application_id)

                # ===== 診斷輸出（debug，確認後可移除）=====
                st.write(f"🔍 DEBUG: matched_row = {'找到了' if matched_row else 'None'}")
                if matched_row:
                    st.write(f"🔍 DEBUG: matched_row keys = {list(matched_row.keys())}")
                # ===== 診斷輸出結束 =====

                if matched_row is None and application_id:
                    st.warning(f"⚠️ Google Sheet 中找不到申請單編號：{application_id}（病史將顯示為未提供）")

                personal_history, family_history = extract_medical_histories(matched_row)
                lifestyle_habits = extract_lifestyle_habits(matched_row)

                smoking_status = lifestyle_habits.get("smoking", "")
                drinking_status = lifestyle_habits.get("drinking", "")
                betel_nut_status = lifestyle_habits.get("betel_nut", "")

                # ===== 診斷輸出（debug，確認後可移除）=====
                st.write(f"🔍 DEBUG: personal_history = '{personal_history}'")
                st.write(f"🔍 DEBUG: family_history = '{family_history}'")
                # ===== 診斷輸出結束 =====

                personal_history = personal_history or "未提供"
                family_history = family_history or ""
                smoking_status = smoking_status or ""
                drinking_status = drinking_status or ""
                betel_nut_status = betel_nut_status or ""
                has_family_history = bool(family_history)
                st.caption(f"檔名：{up_excel.name}｜申請單編號：{application_id or '（無法解析）'}")
                st.caption(f"Google Sheet：{GOOGLE_SHEET_URL}")
                habit_display_parts = []
                if smoking_status:
                    habit_display_parts.append(f"抽菸：{smoking_status}")
                if drinking_status:
                    habit_display_parts.append(f"喝酒：{drinking_status}")
                if betel_nut_status:
                    habit_display_parts.append(f"吃檳榔：{betel_nut_status}")
                habit_display = "｜".join(habit_display_parts) if habit_display_parts else "（未提供）"
                family_display = family_history if has_family_history else "（不參考）"
                st.info(f"個人疾病史：{personal_history}｜家族疾病史：{family_display}｜生活習慣：{habit_display}")

                if not items:
                    st.warning("該檔案中無符合篩選條件的低分項目。")
                else:
                    st.info(f"偵測模式：{mode} | 項目總數：{len(items)}")
                
                final_text = ""
                progress_bar = st.progress(0)
                HEADERS = {
                    "繁體中文": {
                        "intro": "您的檢測結果【{item}】預防評分為低分。",
                        "maintenance": "■ 細胞維護：",
                        "tracking": "■ 主要追蹤項目：",
                        "nutrition": "■ 細胞營養：",
                        "supplements": "■ 功能性營養群建議：",
                        "lifestyle": "■ 生活策略小提醒：",
                    },
                    "English": {
                        "intro": "Your result for 【{item}】 is a low prevention score.",
                        "maintenance": "■ Cellular maintenance:",
                        "tracking": "■ Key tracking labs:",
                        "nutrition": "■ Cellular nutrition:",
                        "supplements": "■ Functional nutrients & supplements:",
                        "lifestyle": "■ Lifestyle tips:",
                    },
                    "日本語": {
                        "intro": "検査結果【{item}】は低スコアです。",
                        "maintenance": "■ 細胞メンテナンス：",
                        "tracking": "■ 追跡すべき検査項目：",
                        "nutrition": "■ 細胞栄養：",
                        "supplements": "■ 栄養補助（サプリ）提案：",
                        "lifestyle": "■ 生活習慣のヒント：",
                    },
                    "한국어": {
                        "intro": "검사 결과【{item}】의 예방 점수가 낮습니다.",
                        "maintenance": "■ 세포 유지:",
                        "tracking": "■ 주요 추적 항목:",
                        "nutrition": "■ 세포 영양:",
                        "supplements": "■ 기능성 영양소/보충제 제안:",
                        "lifestyle": "■ 생활 전략 팁:",
                    },
                    "Tiếng Việt": {
                        "intro": "Kết quả kiểm tra【{item}】có điểm phòng ngừa thấp.",
                        "maintenance": "■ Duy trì tế bào:",
                        "tracking": "■ Các chỉ số cần theo dõi:",
                        "nutrition": "■ Dinh dưỡng tế bào:",
                        "supplements": "■ Gợi ý dưỡng chất/bổ sung:",
                        "lifestyle": "■ Mẹo lối sống:",
                    },
                }
                H = HEADERS.get(lang, HEADERS["繁體中文"])

                # 核心：將 AI 呼叫移入迴圈內，確保每一項都分析到
                for index, item in enumerate(items):
                    st.write(f"正在分析第 {index+1}/{len(items)} 項：{item}...")
                    
                    pdf_tests = "RBC, Hgb, Hct, MCV, MCH, MCHC, Platelet, WBC, Neutrophil, Lymphocyte, Monocyte, Eosinophil, Basophil, Cholesterol, HDL-Cho, LDL-Cho, Triglyceride, Glucose(Fasting/2hrPC), HbA1c, T-Bilirubin, D-Bilirubin, Total Protein, Albumin, Globulin, sGOT, sGPT, Alk-P, r-GTP, BUN, Creatinine, UA, eGFR, AFP, CEA, CA-199, CA-125, CA-153, PSA, CA-724, NSE, cyfra 21-1, SCC, LDH, CPK, HsCRP, Homocysteine, T4, T3, TSH, Free T4, Na, K, Cl, Ca, Phosphorus, EBVCA-IgA, RA, CRP, H. Pylori Ab"
                    generation_limit = max(1, int(word_limit))
                    budget_hint = format_budget_hint(build_length_budget(generation_limit))
                    section_min = min_section_length(word_limit)
                    
                    family_history_instruction_zh = (
                        f"家族疾病史：{family_history}。" if has_family_history else "家族疾病史：不參考。"
                    )
                    family_history_instruction_en = (
                        f"- Family Medical History: {family_history}" if has_family_history else "- Family Medical History: N/A (do not reference family history)"
                    )

                    habit_lines_zh = []
                    habit_lines_en = []
                    if smoking_status:
                        habit_lines_zh.append(f"抽菸問卷結果：{smoking_status}。")
                        habit_lines_en.append(f"- Smoking questionnaire result: {smoking_status}")
                    if drinking_status:
                        habit_lines_zh.append(f"喝酒問卷結果：{drinking_status}。")
                        habit_lines_en.append(f"- Alcohol questionnaire result: {drinking_status}")
                    if betel_nut_status:
                        habit_lines_zh.append(f"吃檳榔問卷結果：{betel_nut_status}。")
                        habit_lines_en.append(f"- Betel nut questionnaire result: {betel_nut_status}")
                    habit_instruction_zh = "\n                    ".join(habit_lines_zh) if habit_lines_zh else ""
                    habit_instruction_en = "\n                    ".join(habit_lines_en) if habit_lines_en else ""
                    smoking_prompt_value = smoking_status or "N/A"
                    drinking_prompt_value = drinking_status or "N/A"
                    betel_prompt_value = betel_nut_status or "N/A"

                    # 強化語言要求，確保 AI 看到
                    user_instruction = f"""
                    ### IMPORTANT LANGUAGE REQUIREMENT: 
                    All content in the JSON response MUST be written in {lang}. 
                    (目前的語言要求：{lang})

                    受試者資料：{user_info.get('gender')}/{user_info.get('age')}歲。
                    申請單編號：{application_id}。
                    個人疾病史：{personal_history}。
                    {family_history_instruction_zh}
                    {habit_instruction_zh}
                    分析項目：{item}。
                    字數限制：{word_limit} 字（以非空白字元計算，請先規劃字數，再產生內容）。
                    生成目標字數：{generation_limit} 字內（需低於或等於字數限制）。
                    各段落字數上限：{budget_hint}。
                    各段落最少字數：{section_min} 字（非空白字元），每段至少 2 句。
                    【追蹤項目】：僅限挑選：[{pdf_tests}]。
                    
                    請嚴格回傳 JSON 格式：
                    {{
                      "maintenance": "...",
                      "tracking": "...",
                      "nutrition": "...",
                      "supplements": "...",
                      "lifestyle": "..."
                    }}
                    """
                    
                    task_prompt = f"""
                    # LANGUAGE CONSTRAINT (CRITICAL)
                    - YOU MUST RESPOND EXCLUSIVELY IN: {lang}
                    - IF {lang} IS "English", DO NOT USE ANY CHINESE CHARACTERS.
                    - IF {lang} IS "日本語", すべて日本語で回答してください。
                    - IF {lang} IS "한국어", 한국어로만 작성하세요.
                    - IF {lang} IS "Tiếng Việt", chỉ trả lời bằng tiếng Việt.

                    # SUBJECT DATA
                    - Gender/Age: {user_info.get('gender')}/{user_info.get('age')}
                    - Application ID: {application_id}
                    - Personal Medical History: {personal_history}
                    {family_history_instruction_en}
                    {habit_instruction_en}
                    - Smoking Status (binary): {smoking_prompt_value}
                    - Alcohol Status (binary): {drinking_prompt_value}
                    - Betel Nut Status (binary): {betel_prompt_value}
                    - Target Item: {item}
                    - Word Limit (Hard Max, non-space characters): {word_limit}
                    - Target Limit (Use This): {generation_limit}
                    - Section Budgets: {budget_hint}
                    - Minimum Per Section: {section_min} (non-space characters), at least 2 sentences each

                    # REFERENCE DATA (FOR TRACKING SECTION)
                    - Valid Tracking Items: [{pdf_tests}]

                    # RESPONSE FORMAT
                    - STRICT: If family history is marked as N/A, do not mention family history at all.
                    - STRICT: Mention smoking/alcohol/betel nut ONLY when the corresponding status is 「有」.
                    - STRICT: If a habit is 「無」, "N/A", or empty, DO NOT provide related risk claims or lifestyle advice for that habit. 
                    - STRICT: Use only disease-to-gene mappings explicitly defined in the system prompt; do not invent or substitute genes.
                    - IF the target item has no explicit gene mapping in the system prompt, avoid naming any gene.
                    - Focus on mechanisms strictly relevant to the target item.
                    Please provide the analysis strictly in the following JSON structure:
                    {{
                    "maintenance": "...",
                    "tracking": "...",
                    "nutrition": "...",
                    "supplements": "...",
                    "lifestyle": "..."
                    }}
                    """

                    lifestyle_guidance = """
                    # LIFESTYLE GUIDANCE (TOPIC-ALIGNED, QUANTIFIABLE)
                    Provide 3-6 actionable lifestyle tips tailored to the user's age/gender and the target item.
                    Every tip must be measurable (frequency, duration, timing, or quantity).
                    Ensure each tip is explicitly connected to the target topic's mechanism.
                    Avoid vague or non-quantifiable items (e.g., meditation, deep breathing, "sleep early").
                    Each section must include at least 2 sentences and avoid empty headers.
                    """

                    # 2. 使用 system_instruction 分離角色與任務
                    system_prompt = bg_prompt + "\n\n" + build_language_system_rule(lang, generation_limit)
                    full_combined_prompt = f"{system_prompt}\n\n{user_instruction}\n\n{task_prompt}\n\n{lifestyle_guidance}"
                    report = None
                    failure_reason = ""
                    output_length = 0
                    for attempt in range(3):
                        if attempt == 1:
                            if output_length > word_limit:
                                shrink_by = max(10, output_length - word_limit)
                                generation_limit = max(1, generation_limit - shrink_by)
                            budget_hint = format_budget_hint(build_length_budget(generation_limit))
                            section_min = min_section_length(word_limit)
                            system_prompt = bg_prompt + "\n\n" + build_language_system_rule(lang, generation_limit)
                            user_instruction = f"""
                            ### IMPORTANT LANGUAGE REQUIREMENT: 
                            All content in the JSON response MUST be written in {lang}. 
                            (目前的語言要求：{lang})

                            受試者資料：{user_info.get('gender')}/{user_info.get('age')}歲。
                            申請單編號：{application_id}。
                            個人疾病史：{personal_history}。
                            {family_history_instruction_zh}
                            {habit_instruction_zh}
                            分析項目：{item}。
                            字數限制：{word_limit} 字（以非空白字元計算，請先規劃字數，再產生內容）。
                            生成目標字數：{generation_limit} 字內（需低於或等於字數限制）。
                            各段落字數上限：{budget_hint}。
                            各段落最少字數：{section_min} 字（非空白字元），每段至少 2 句。
                            【追蹤項目】：僅限挑選：[{pdf_tests}]。
                            
                            請嚴格回傳 JSON 格式：
                            {{
                              "maintenance": "...",
                              "tracking": "...",
                              "nutrition": "...",
                              "supplements": "...",
                              "lifestyle": "..."
                            }}
                            """
                            task_prompt = f"""
                            # LANGUAGE CONSTRAINT (CRITICAL)
                            - YOU MUST RESPOND EXCLUSIVELY IN: {lang}
                            - IF {lang} IS "English", DO NOT USE ANY CHINESE CHARACTERS.
                            - IF {lang} IS "日本語", すべて日本語で回答してください。
                            - IF {lang} IS "한국어", 한국어로만 작성하세요.
                            - IF {lang} IS "Tiếng Việt", chỉ trả lời bằng tiếng Việt.

                            # SUBJECT DATA
                            - Gender/Age: {user_info.get('gender')}/{user_info.get('age')}
                            - Application ID: {application_id}
                            - Personal Medical History: {personal_history}
                            {family_history_instruction_en}
                            {habit_instruction_en}
                            - Smoking Status (binary): {smoking_prompt_value}
                            - Alcohol Status (binary): {drinking_prompt_value}
                            - Betel Nut Status (binary): {betel_prompt_value}
                            - Target Item: {item}
                            - Word Limit (Hard Max, non-space characters): {word_limit}
                            - Target Limit (Use This): {generation_limit}
                            - Section Budgets: {budget_hint}
                            - Minimum Per Section: {section_min} (non-space characters), at least 2 sentences each

                            # REFERENCE DATA (FOR TRACKING SECTION)
                            - Valid Tracking Items: [{pdf_tests}]

                            # RESPONSE FORMAT
                            - STRICT: If family history is marked as N/A, do not mention family history at all.
                            - STRICT: Mention smoking/alcohol/betel nut ONLY when the corresponding status is 「有」.
                            - STRICT: If a habit is 「無」, "N/A", or empty, DO NOT provide related risk claims or lifestyle advice for that habit. 
                            - STRICT: Use only disease-to-gene mappings explicitly defined in the system prompt; do not invent or substitute genes.
                            - IF the target item has no explicit gene mapping in the system prompt, avoid naming any gene.
                            - Focus on mechanisms strictly relevant to the target item.
                            Please provide the analysis strictly in the following JSON structure:
                            {{
                            "maintenance": "...",
                            "tracking": "...",
                            "nutrition": "...",
                            "supplements": "...",
                            "lifestyle": "..."
                            }}
                            """
                            full_combined_prompt = f"{system_prompt}\n\n{user_instruction}\n\n{task_prompt}\n\n{lifestyle_guidance}"
                            full_combined_prompt += (
                                f"\n\n# RETRY NOTICE\n"
                                f"The previous response was invalid: {failure_reason}.\n"
                                f"Please respond again strictly in {lang} and within the target limit.\n"
                            )
                        response = client.models.generate_content(
                            model="models/gemma-3-27b-it",
                            contents=full_combined_prompt,
                            config={
                                "temperature": 0.3,
                                "top_p": 0.95,
                            }
                        )

                        json_match = re.search(r'\{.*\}', response.text, re.DOTALL)
                        if not json_match:
                            failure_reason = "未回傳有效 JSON"
                            continue

                        candidate_report = json.loads(json_match.group(0))
                        valid, failure_reason, output_length = validate_report_output(candidate_report, lang, word_limit)
                        if valid:
                            report = candidate_report
                            break

                    if report:
                        section = H["intro"].format(item=item) + "\n\n"
                        section += f'{H["maintenance"]}\n{format_output(report.get("maintenance"))}\n\n'
                        section += f'{H["tracking"]}\n{format_output(report.get("tracking"))}\n\n'
                        section += f'{H["nutrition"]}\n{format_output(report.get("nutrition"))}\n\n'
                        section += f'{H["supplements"]}\n{format_output(report.get("supplements"))}\n\n'
                        section += f'{H["lifestyle"]}\n{format_output(report.get("lifestyle"))}\n\n'
                        final_text += section + "="*50 + "\n\n"
                    else:
                        st.warning(f"第 {index+1} 項分析失敗：{failure_reason}")
                    
                    progress_bar.progress((index + 1) / len(items))
                    if len(items) > 1:
                        time.sleep(5) # 避免頻率限制

                st.success("🎉 分析完成！")
                st.text_area("結果預覽", final_text, height=400)
                st.download_button("📥 下載報告", final_text, file_name="分析報告.txt")

        except Exception as e:
            st.error(f"分析失敗：{e}")
